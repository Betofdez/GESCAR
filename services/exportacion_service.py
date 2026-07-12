from io import BytesIO
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


def convertir_valor_excel(valor):
    """
    Convierte valores Decimal a float para que Excel pueda tratarlos
    como números y permita sumarlos, ordenarlos o aplicar fórmulas.
    """
    if isinstance(valor, Decimal):
        return float(valor)

    return valor


def formatear_valor_pdf(valor):
    """
    Convierte los números al formato español para mostrarlos en PDF.
    """
    if isinstance(valor, (int, float, Decimal)):
        texto = f"{float(valor):,.2f}"
        return texto.replace(",", "TEMP").replace(".", ",").replace("TEMP", ".")

    if valor is None:
        return ""

    return str(valor)


def crear_excel(nombre_hoja, columnas, filas):
    """
    Genera un fichero Excel en memoria.

    columnas:
        Lista con los títulos de las columnas.

    filas:
        Lista de listas con los datos.
    """
    libro = Workbook()
    hoja = libro.active
    hoja.title = nombre_hoja[:31]

    color_cabecera = "003366"

    # Crear cabecera
    for numero_columna, titulo in enumerate(columnas, start=1):
        celda = hoja.cell(row=1, column=numero_columna, value=titulo)

        celda.font = Font(
            bold=True,
            color="FFFFFF"
        )

        celda.fill = PatternFill(
            fill_type="solid",
            fgColor=color_cabecera
        )

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # Insertar datos
    for numero_fila, fila in enumerate(filas, start=2):
        for numero_columna, valor in enumerate(fila, start=1):
            valor_convertido = convertir_valor_excel(valor)

            celda = hoja.cell(
                row=numero_fila,
                column=numero_columna,
                value=valor_convertido
            )

            celda.alignment = Alignment(
                vertical="center"
            )

            if isinstance(valor_convertido, (int, float)):
                celda.number_format = '#,##0.00'

    # Ajustar automáticamente el ancho de las columnas
    for columna in hoja.columns:
        longitud_maxima = 0
        letra_columna = get_column_letter(columna[0].column)

        for celda in columna:
            if celda.value is not None:
                longitud = len(str(celda.value))
                longitud_maxima = max(longitud_maxima, longitud)

        hoja.column_dimensions[letra_columna].width = min(
            longitud_maxima + 3,
            35
        )

    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions

    archivo = BytesIO()
    libro.save(archivo)
    archivo.seek(0)

    return archivo


def crear_pdf(titulo, columnas, filas):
    """
    Genera un PDF horizontal en memoria con una tabla.
    """
    archivo = BytesIO()

    documento = SimpleDocTemplate(
        archivo,
        pagesize=landscape(A4),
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=1 * cm,
        bottomMargin=1 * cm
    )

    estilos = getSampleStyleSheet()
    elementos = []

    titulo_pdf = Paragraph(titulo, estilos["Title"])
    elementos.append(titulo_pdf)
    elementos.append(Spacer(1, 0.5 * cm))

    datos_tabla = [columnas]

    for fila in filas:
        datos_tabla.append([
            formatear_valor_pdf(valor)
            for valor in fila
        ])

    tabla = Table(
        datos_tabla,
        repeatRows=1
    )

    tabla.setStyle(TableStyle([
        (
            "BACKGROUND",
            (0, 0),
            (-1, 0),
            colors.HexColor("#003366")
        ),
        (
            "TEXTCOLOR",
            (0, 0),
            (-1, 0),
            colors.white
        ),
        (
            "FONTNAME",
            (0, 0),
            (-1, 0),
            "Helvetica-Bold"
        ),
        (
            "ALIGN",
            (0, 0),
            (-1, 0),
            "CENTER"
        ),
        (
            "VALIGN",
            (0, 0),
            (-1, -1),
            "MIDDLE"
        ),
        (
            "GRID",
            (0, 0),
            (-1, -1),
            0.5,
            colors.grey
        ),
        (
            "ROWBACKGROUNDS",
            (0, 1),
            (-1, -1),
            [
                colors.white,
                colors.HexColor("#F2F4F7")
            ]
        ),
        (
            "FONTSIZE",
            (0, 0),
            (-1, -1),
            7
        ),
        (
            "LEFTPADDING",
            (0, 0),
            (-1, -1),
            4
        ),
        (
            "RIGHTPADDING",
            (0, 0),
            (-1, -1),
            4
        )
    ]))

    elementos.append(tabla)
    documento.build(elementos)

    archivo.seek(0)

    return archivo